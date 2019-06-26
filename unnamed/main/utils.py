from django.http import HttpResponse
from openpyxl import Workbook


def write_to_excel(fields, q):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'data'

    columns = [f for f in fields['purchase']]
    columns.extend(['customer_' + f for f in fields['customer']])
    columns.extend(['item_' + f for f in fields['item']])

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for p in q:
        row_num += 1

        row = [
            getattr(p, field) if field not in ('customer', 'item')
            else getattr(p, field).id
            for field in fields['purchase']
        ]
        row.extend([
            getattr(getattr(p, key), field)
            for key in fields if key != "purchase"
            for field in fields[key]
        ])

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)

    workbook.save(response)
    return response
